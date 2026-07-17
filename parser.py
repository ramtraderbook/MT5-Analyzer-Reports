"""
parser.py - MT5 xlsx trade history parser for EA Analyzer
Detects sections dynamically, joins POSITIONS with ORDERS to get EA comments.
"""

import math
import re
import openpyxl
from datetime import datetime, date
from decimal import Decimal


SECTION_KEYWORDS = {"Positions", "Orders", "Deals", "Open Positions", "Results"}

# MT5 auto-generates comments starting with '[' for SL/TP/SO/OUT/expiration
# triggers (e.g. "[sl 1234.5]", "[tp 1234.5]", "[so: ...]", "[out 1234.5]",
# "[expiration]"). Kept for backward compatibility with callers/tests that
# only need the literal prefix character; the actual filtering logic now
# uses the more precise _is_system_comment() below, since a legit EA name
# like "[GridMaster v2]" also starts with '[' but is NOT a system comment.
SYSTEM_COMMENT_PREFIX = "["

_SYSTEM_COMMENT_RE = re.compile(r"^\[(sl|tp|so|out|expiration)\b", re.IGNORECASE)


def _is_system_comment(s: str) -> bool:
    """True if s is an MT5 auto-generated system comment (SL/TP/SO/OUT/expiration triggers)."""
    return bool(_SYSTEM_COMMENT_RE.match(s))


def _is_numeric_cell(v) -> bool:
    """
    True if v is a numeric value: a native int/float/Decimal (bool
    excluded), or a string that parses as a number. Used by the
    OPEN POSITIONS comment fallback scan to skip stray numeric cells
    (e.g. Profit/Swap) instead of misreading them as an EA name.
    """
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float, Decimal)):
        return True
    if isinstance(v, str):
        try:
            float(v.strip())
            return True
        except ValueError:
            return False
    return False


_DATE_FORMATS = (
    "%Y.%m.%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%Y.%m.%d %H:%M",
    "%Y-%m-%d %H:%M",
    "%Y.%m.%d",
    "%Y-%m-%d",
    "%Y.%m.%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S.%f",
)


def _parse_date(val):
    """
    Parse an MT5 date cell (string, datetime, or date) into a datetime.

    Accepts native datetime objects as-is and datetime.date objects
    (converted to midnight datetime), then tries a series of known
    MT5/ISO string formats (with and without seconds/fractional seconds),
    then datetime.fromisoformat as a final fallback.

    Deliberately does NOT fall back to a day-first "%d.%m.%Y" format: MT5
    never emits day-first exports, and guessing day-first would silently
    transpose month/day on a month-first "MM.DD.YYYY" export instead of
    failing loudly.
    Returns None for genuinely unparseable input.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)

    s = str(val).strip()
    if not s:
        return None

    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


_CURRENCY_SYMBOLS = "$€£¥"
_CURRENCY_CODE_RE = re.compile(r"(?i)\b(USD|EUR|GBP|JPY|CHF|AUD|CAD|NZD|CNY)\b")
_THOUSANDS_ONLY_COMMA_RE = re.compile(r"^-?\d{1,3}(,\d{3})+$")
_ALL_ZEROS_RE = re.compile(r"^-?0+$")


def _to_float(val, default=0.0, ambiguous_comma="thousands"):
    """
    Safely convert an MT5 cell value to float.

    Handles:
      - native int/float/Decimal passthrough (no string round-trip);
        bool is explicitly excluded (bool is a subclass of int in
        Python, but a checkbox/boolean cell is not a numeric value)
      - currency symbols ($ EUR GBP JPY CHF...) and alpha currency codes
      - parenthesized negatives: "(12.34)" -> -12.34
      - mixed thousands/decimal separators: when BOTH ',' and '.' are
        present, the LAST-occurring one is the decimal separator
      - comma-only values: see the ambiguity note below
      - space / NBSP used as thousands grouping
      - non-finite results (nan/inf/-inf, whether parsed from a string
        like "nan"/"inf"/"Infinity" or arriving as a native float from
        openpyxl) are rejected -> `default`, never returned as a value

    Comma-only ambiguity (X,YYY shape):
      A comma-only value like "1,234" is GENUINELY AMBIGUOUS from the
      string alone: it could be thousands grouping (1234.0) or an EU
      decimal comma with exactly 3 fractional digits (1.234). This is
      NOT resolvable per-cell — it depends on the column's semantics,
      which the caller knows and the string does not. Two sub-cases:

      - If the group before the first (and only) comma is all zeros
        ("0", "-0", "00", ...), it can NEVER be thousands grouping (no
        real number is thousands-grouped as "0,NNN") — always treated as
        a decimal comma regardless of `ambiguous_comma` (e.g. "0,001" ->
        0.001, "0,500" -> 0.5, "00,001" -> 0.001).
      - If there is more than one comma group (e.g. "1,234,567" or
        "0,001,000"), it is unambiguously thousands grouping — a decimal
        number can only have one decimal separator — even when the
        leading group is all zeros.
      - Otherwise (single nonzero 1-3 digit group + exactly 3 decimal
        digits, e.g. "1,234" / "145,678"), the shape is truly ambiguous
        and `ambiguous_comma` decides:
          * "thousands" (default): treated as thousands -> "1,000" -> 1000.0.
            Use for MONEY columns (commission/swap/profit).
          * "decimal": treated as an EU decimal comma -> "145,678" -> 145.678.
            Use for PRICE/QUANTITY columns (volume/open_price/close_price/sl/tp).
    Unparseable input returns `default` — but `default` itself is validated
    the same way as any computed result: a non-finite `default` (e.g.
    float("nan")) is never returned either, and is normalized to 0.0 once
    at function entry so every `return default` path stays finite.
    """
    if (
        isinstance(default, (int, float, Decimal))
        and not isinstance(default, bool)
        and not math.isfinite(default)
    ):
        default = 0.0

    if val is None:
        return default
    if isinstance(val, bool):
        return default
    if isinstance(val, (int, float, Decimal)):
        result = float(val)
        return result if math.isfinite(result) else default

    s = str(val).strip()
    if not s:
        return default

    # Normalize NBSP / narrow NBSP (sometimes used as thousands grouping)
    s = s.replace(" ", "").replace(" ", "")

    # Parenthesized negative: "(12.34)" -> negative
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    # Strip currency symbols
    for ch in _CURRENCY_SYMBOLS:
        s = s.replace(ch, "")

    # Strip leading/trailing alpha currency codes ("USD 12.34" / "12.34 USD")
    s = _CURRENCY_CODE_RE.sub("", s).strip()

    if not s:
        return default

    has_comma = "," in s
    has_dot = "." in s

    if has_comma and has_dot:
        # Whichever separator occurs LAST is the decimal separator
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif has_comma:
        first_group = s.split(",", 1)[0]
        if s.count(",") == 1 and _ALL_ZEROS_RE.match(first_group):
            # Rule (a): a single-group all-zeros leading group is never
            # thousands grouping — no real number is thousands-grouped as
            # "0,NNN". Restricted to the single-comma case: a multi-group
            # value with an all-zero leading group (e.g. "0,001,000") is
            # still unambiguously thousands (see the multi-group branch
            # below) and must not be short-circuited here.
            s = s.replace(",", ".")
        elif _THOUSANDS_ONLY_COMMA_RE.match(s):
            if s.count(",") > 1:
                # Multiple thousands groups: unambiguously thousands
                # (a decimal number can only have one decimal separator).
                s = s.replace(",", "")
            else:
                # Rule (b): single nonzero 1-3 digit group + exactly 3
                # decimal digits — genuinely ambiguous from the string
                # alone; resolved by the caller's column semantics.
                if ambiguous_comma == "decimal":
                    s = s.replace(",", ".")
                else:
                    s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    # else: only '.' or no separator at all -> leave as-is

    # Remove spaces used as thousands grouping (e.g. "1 234.56")
    s = s.replace(" ", "")

    try:
        result = float(s)
    except (ValueError, TypeError):
        return default

    if not math.isfinite(result):
        return default

    if negative:
        result = -abs(result)
    return result


def _to_price_or_none(val):
    """
    Parse an MT5 S/L or T/P cell to a positive price, or None if unset.

    In MT5, S/L = 0 (or T/P = 0) means "no stop/target was set" — it is a
    sentinel for absence, not a price of zero. That sentinel arrives from
    the export in several shapes: native 0.0, the string "0"/"0.0"/"0,0",
    an empty cell, or None. All of them must map to None so downstream
    code can never mistake an unset level for a real level priced at zero.

    The previous per-call `if sl_val` guard was inconsistent: it tested
    the RAW cell, so numeric 0.0 (falsy) became None but the string "0"
    (truthy) parsed to 0.0 and was treated as a real price. Parsing first,
    then keeping only strictly-positive results, removes that split — a
    valid price is always > 0, so anything <= 0 (including negatives from
    a malformed cell) is treated as unset.

    Uses ambiguous_comma="decimal": S/L and T/P are price columns.
    """
    price = _to_float(val, default=0.0, ambiguous_comma="decimal")
    return price if price > 0 else None


def _find_section_rows(ws):
    """
    Scan column A for section header keywords.
    Returns dict: { "Positions": row_idx, "Orders": row_idx, ... }

    Keeps the FIRST occurrence of each keyword (the real section header).
    A later duplicate of the same keyword in column A (e.g. inside a data
    row) is ignored instead of silently overwriting the real section start.
    """
    sections = {}
    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if isinstance(cell_val, str):
            key = cell_val.strip()
            if key in SECTION_KEYWORDS and key not in sections:
                sections[key] = row_idx
    return sections


def _compute_boundaries(section_starts, max_row):
    """
    Returns list of (name, header_row, data_start_row, end_row) tuples.
    header_row = section_row + 1 (column headers)
    data_start_row = header_row + 1 (first data row)
    end_row = next section start - 1
    """
    sorted_sections = sorted(section_starts.items(), key=lambda x: x[1])
    boundaries = []
    for i, (name, section_row) in enumerate(sorted_sections):
        header_row = section_row + 1
        data_start = section_row + 2
        if i + 1 < len(sorted_sections):
            end_row = sorted_sections[i + 1][1] - 1
        else:
            end_row = max_row
        boundaries.append((name, header_row, data_start, end_row))
    return boundaries


def _get_column_map(ws, header_row):
    """
    Read header row and return {column_name: column_index (1-based)}.
    On duplicate header names, the LAST occurrence wins. Kept for callers
    that only care about unique columns; see _get_column_map_all() for
    columns (like "Time"/"Price" in POSITIONS) that repeat.
    """
    col_map = {}
    for col in range(1, 30):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            col_map[str(val).strip()] = col
    return col_map


def _get_column_map_all(ws, header_row):
    """
    Read header row and return {column_name: [col_idx, ...]}, preserving
    ALL occurrences in left-to-right order. Needed for POSITIONS, where
    "Time" and "Price" each appear twice (open + close).
    """
    col_map = {}
    for col in range(1, 30):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            col_map.setdefault(str(val).strip(), []).append(col)
    return col_map


def _parse_header(ws):
    """Extract account info from the first 5 rows."""
    account = {
        "name": "",
        "number": "",
        "currency": "USD",
        "broker": "",
        "report_date": ""
    }
    for row_idx in range(1, 6):
        val = ws.cell(row=row_idx, column=1).value
        if not val:
            continue
        s = str(val)
        if s.startswith("Name:"):
            account["name"] = s.replace("Name:", "").strip()
        elif s.startswith("Account:"):
            parts = s.replace("Account:", "").strip()
            # Extract account number (first token)
            tokens = parts.split()
            if tokens:
                account["number"] = tokens[0].strip("()")
            # Extract currency
            for t in tokens:
                if t in ("USD", "EUR", "GBP"):
                    account["currency"] = t
                    break
        elif s.startswith("Company:"):
            account["broker"] = s.replace("Company:", "").strip()
        elif s.startswith("Date:"):
            account["report_date"] = s.replace("Date:", "").strip()
    return account


def _parse_positions(ws, header_row, data_start, end_row):
    """
    Parse POSITIONS section. Returns list of trade dicts.
    Columns: Time(open), Position, Symbol, Type, Volume, Price(open),
             S/L, T/P, Time(close), Price(close), Commission, Swap, Profit

    NOTE: "Time" and "Price" appear TWICE in headers (open + close).
    _get_column_map() only stores the LAST occurrence, so duplicate
    columns are resolved via _get_column_map_all() by OCCURRENCE ORDER
    in the real header row (1st "Time"/"Price" = open, 2nd = close).
    This is only a positional fallback if the header genuinely lacks
    that many occurrences (e.g. a truncated/malformed report).
    """
    col_map = _get_column_map(ws, header_row)
    col_map_all = _get_column_map_all(ws, header_row)

    # For unique column names, use col_map with positional fallback
    def get_col(name, fallback):
        return col_map.get(name, fallback)

    def get_nth_col(name, occurrence_index, fallback):
        """Resolve the Nth (0-based) occurrence of a duplicate header name."""
        cols = col_map_all.get(name, [])
        if len(cols) > occurrence_index:
            return cols[occurrence_index]
        return fallback

    # Duplicate header names (Time, Price appear twice) → resolve by order
    time_open_col   = get_nth_col("Time", 0, 1)    # 1st "Time" → open
    price_open_col  = get_nth_col("Price", 0, 6)   # 1st "Price" → open
    time_close_col  = get_nth_col("Time", 1, 9)    # 2nd "Time" → close
    price_close_col = get_nth_col("Price", 1, 10)  # 2nd "Price" → close

    # Unique column names — safe to use col_map
    position_col   = get_col("Position", 2)
    symbol_col     = get_col("Symbol", 3)
    type_col       = get_col("Type", 4)
    volume_col     = get_col("Volume", 5)
    sl_col         = col_map.get("S / L", col_map.get("S/L", 7))
    tp_col         = col_map.get("T / P", col_map.get("T/P", 8))
    commission_col = get_col("Commission", 11)
    swap_col       = get_col("Swap", 12)
    profit_col     = get_col("Profit", 13)

    trades = []
    for row_idx in range(data_start, end_row + 1):
        pos_id_val = ws.cell(row=row_idx, column=position_col).value
        if pos_id_val is None:
            continue

        # Skip rows that are subtotals or headers (non-numeric position ID)
        try:
            pos_id = int(pos_id_val)
        except (ValueError, TypeError):
            continue

        open_time = _parse_date(ws.cell(row=row_idx, column=time_open_col).value)
        close_time = _parse_date(ws.cell(row=row_idx, column=time_close_col).value)
        commission = _to_float(ws.cell(row=row_idx, column=commission_col).value)
        swap = _to_float(ws.cell(row=row_idx, column=swap_col).value)
        profit = _to_float(ws.cell(row=row_idx, column=profit_col).value)
        net_pnl = profit + commission + swap

        sl_val = ws.cell(row=row_idx, column=sl_col).value
        tp_val = ws.cell(row=row_idx, column=tp_col).value

        duration_hours = 0.0
        if open_time and close_time:
            delta = close_time - open_time
            duration_hours = delta.total_seconds() / 3600

        trades.append({
            "position_id": pos_id,
            "symbol": str(ws.cell(row=row_idx, column=symbol_col).value or "").strip(),
            "direction": str(ws.cell(row=row_idx, column=type_col).value or "").strip().lower(),
            "volume": _to_float(ws.cell(row=row_idx, column=volume_col).value, ambiguous_comma="decimal"),
            "open_time": open_time,
            "close_time": close_time,
            "open_price": _to_float(ws.cell(row=row_idx, column=price_open_col).value, ambiguous_comma="decimal"),
            "close_price": _to_float(ws.cell(row=row_idx, column=price_close_col).value, ambiguous_comma="decimal"),
            "sl": _to_price_or_none(sl_val),
            "tp": _to_price_or_none(tp_val),
            "commission": commission,
            "swap": swap,
            "profit": profit,
            "net_pnl": net_pnl,
            "duration_hours": round(duration_hours, 2),
            "comment": "Unknown"  # will be filled by JOIN with ORDERS
        })

    return trades


def _parse_orders(ws, header_row, data_start, end_row):
    """
    Parse ORDERS section. Returns {order_id (int): comment (str)}.
    Only keeps rows with a non-empty comment that is not an MT5 system
    comment (see _is_system_comment). On duplicate order_id rows, the
    FIRST non-system comment found is kept — a later row for the same
    order_id no longer silently overwrites it.
    """
    col_map = _get_column_map(ws, header_row)

    order_col = col_map.get("Order", 2)
    comment_col = col_map.get("Comment", 12)

    order_map = {}
    for row_idx in range(data_start, end_row + 1):
        order_val = ws.cell(row=row_idx, column=order_col).value
        comment_val = ws.cell(row=row_idx, column=comment_col).value

        if order_val is None:
            continue

        try:
            order_id = int(order_val)
        except (ValueError, TypeError):
            continue

        comment_str = str(comment_val or "").strip()
        # Keep any non-empty comment that isn't an MT5 system comment (e.g. "[sl 1234.5]")
        if comment_str and not _is_system_comment(comment_str) and order_id not in order_map:
            order_map[order_id] = comment_str

    return order_map


def _parse_open_positions(ws, header_row, data_start, end_row):
    """
    Parse OPEN POSITIONS section.

    The standard MT5 Open Positions layout has NO "Commission" column
    (Swap/Profit/Comment follow directly), so commission_col is only
    resolved from a real "Commission" header — never a positional
    fallback onto another column (that would double-count e.g. Swap
    into net_pnl). "Comment" is resolved from the header when present;
    the positional scan is only a last resort when the header lacks it.
    """
    col_map = _get_column_map(ws, header_row)

    time_col = col_map.get("Time", 1)
    position_col = col_map.get("Position", 2)
    symbol_col = col_map.get("Symbol", 3)
    type_col = col_map.get("Type", 4)
    volume_col = col_map.get("Volume", 5)
    price_open_col = col_map.get("Price", 6)
    sl_col = col_map.get("S / L", col_map.get("S/L", 7))
    tp_col = col_map.get("T / P", col_map.get("T/P", 8))
    commission_col = col_map.get("Commission")  # no positional fallback — may legitimately be absent
    swap_col = col_map.get("Swap", 12)
    profit_col = col_map.get("Profit", 13)
    comment_col = col_map.get("Comment")  # resolved by header; fallback scan only if absent

    open_positions = []
    for row_idx in range(data_start, end_row + 1):
        pos_id_val = ws.cell(row=row_idx, column=position_col).value
        if pos_id_val is None:
            continue
        try:
            pos_id = int(pos_id_val)
        except (ValueError, TypeError):
            continue

        if commission_col is not None:
            commission = _to_float(ws.cell(row=row_idx, column=commission_col).value)
        else:
            commission = 0.0
        swap = _to_float(ws.cell(row=row_idx, column=swap_col).value)
        profit = _to_float(ws.cell(row=row_idx, column=profit_col).value)

        if comment_col is not None:
            raw_comment = ws.cell(row=row_idx, column=comment_col).value
            comment_val = None
            if raw_comment and str(raw_comment).strip() and not _is_system_comment(str(raw_comment).strip()):
                comment_val = raw_comment
        else:
            # Header lacks a "Comment" column — scan nearby columns as a
            # last resort. Skip any cell whose value is numeric (int/float,
            # or a string that parses as a number) — accepting it would
            # resurrect a narrower version of the original "EA name became
            # 150.75" bug by grabbing a stray Profit/Swap cell. Only a
            # genuinely non-numeric text value qualifies as an EA name.
            #
            # Known limitation (accepted tradeoff): this also skips a
            # legitimate EA whose comment is a bare magic number (e.g.
            # "1104", which trade_matching.py matches against a mapping's
            # magic) when the export is malformed in this exact way
            # (missing "Comment" header AND a purely-numeric comment
            # cell). That is intentional — a phantom EA named "150.75"
            # silently polluting ea_names for every malformed export is
            # worse than an honest "Unknown" for this narrow, unusual
            # case.
            comment_val = None
            for c in range(13, 17):
                v = ws.cell(row=row_idx, column=c).value
                if v is None or _is_numeric_cell(v):
                    continue
                if str(v).strip() and not _is_system_comment(str(v).strip()):
                    comment_val = v
                    break

        sl_val = ws.cell(row=row_idx, column=sl_col).value
        tp_val = ws.cell(row=row_idx, column=tp_col).value

        open_positions.append({
            "position_id": pos_id,
            "comment": str(comment_val or "Unknown").strip(),
            "symbol": str(ws.cell(row=row_idx, column=symbol_col).value or "").strip(),
            "direction": str(ws.cell(row=row_idx, column=type_col).value or "").strip().lower(),
            "volume": _to_float(ws.cell(row=row_idx, column=volume_col).value, ambiguous_comma="decimal"),
            "open_time": _parse_date(ws.cell(row=row_idx, column=time_col).value),
            "open_price": _to_float(ws.cell(row=row_idx, column=price_open_col).value, ambiguous_comma="decimal"),
            "sl": _to_price_or_none(sl_val),
            "tp": _to_price_or_none(tp_val),
            "commission": commission,
            "swap": swap,
            "profit": profit,
            "net_pnl": profit + commission + swap,
        })

    return open_positions


def _parse_results(ws, header_row, data_start, end_row):
    """
    Parse RESULTS section for validation.
    Returns dict of label -> value pairs.
    """
    results = {}
    for row_idx in range(header_row, end_row + 1):
        for col in range(1, 10):
            label = ws.cell(row=row_idx, column=col).value
            if label and isinstance(label, str) and label.strip():
                val = ws.cell(row=row_idx, column=col + 1).value
                results[label.strip()] = val
                break
    return results


def parse_mt5_report(filepath: str) -> dict:
    """
    Main entry point. Parses an MT5 xlsx trade history report.
    Returns ParsedData dict with account info, closed trades, open positions.

    Raises ValueError with descriptive message on parse failure.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo xlsx: {e}")

    ws = wb.active

    # Find all section boundaries
    section_starts = _find_section_rows(ws)

    required = ["Positions", "Orders"]
    for req in required:
        if req not in section_starts:
            raise ValueError(
                f"Sección '{req}' no encontrada en el archivo. "
                "Verifica que el archivo sea un reporte de MT5 válido."
            )

    boundaries = _compute_boundaries(section_starts, ws.max_row)
    bound_map = {name: (hr, ds, er) for name, hr, ds, er in boundaries}

    # Parse header (account info)
    account = _parse_header(ws)

    # Parse POSITIONS (closed trades without comments)
    pos_hr, pos_ds, pos_er = bound_map["Positions"]
    closed_trades = _parse_positions(ws, pos_hr, pos_ds, pos_er)

    # Parse ORDERS (get EA comment for each position)
    ord_hr, ord_ds, ord_er = bound_map["Orders"]
    order_map = _parse_orders(ws, ord_hr, ord_ds, ord_er)

    # JOIN: enrich each trade with EA comment
    unknown_count = 0
    for trade in closed_trades:
        pid = trade["position_id"]
        if pid in order_map:
            trade["comment"] = order_map[pid]
        else:
            trade["comment"] = "Unknown"
            unknown_count += 1

    # Parse OPEN POSITIONS (if present)
    open_positions = []
    if "Open Positions" in bound_map:
        op_hr, op_ds, op_er = bound_map["Open Positions"]
        open_positions = _parse_open_positions(ws, op_hr, op_ds, op_er)

    # Parse RESULTS (for validation)
    results_validation = {}
    if "Results" in bound_map:
        res_hr, res_ds, res_er = bound_map["Results"]
        results_validation = _parse_results(ws, res_hr, res_ds, res_er)

    wb.close()

    # Build sorted list of unique EA names (exclude Unknown)
    ea_names = sorted(set(
        t["comment"] for t in closed_trades
        if t["comment"] != "Unknown"
    ))

    return {
        "account": account,
        "closed_trades": closed_trades,
        "open_positions": open_positions,
        "ea_names": ea_names,
        "results_validation": results_validation,
        "unknown_trades": unknown_count,
        "total_closed": len(closed_trades),
        "total_open": len(open_positions),
    }


def merge_trades(existing: list, new_trades: list) -> list:
    """
    Merge two lists of trade dicts, deduplicating by position_id.

    New trades take precedence — if the same position_id appears in
    both lists (or more than once within new_trades itself), the trade
    from new_trades REPLACES any earlier one. Within new_trades, the
    LAST occurrence wins. This lets a re-uploaded file carry broker
    corrections (e.g. a commission fixed after the fact) forward into
    the cache, and prevents a single upload that lists the same
    position twice from double-counting P&L.

    Returns a new list sorted by close_time (ascending), stable for
    equal close_time. This is the core of the append-mode multi-file
    loading feature.
    """
    merged_by_id = {t["position_id"]: t for t in existing}
    for t in new_trades:
        merged_by_id[t["position_id"]] = t  # new always replaces existing; last new wins
    merged = list(merged_by_id.values())

    def sort_key(t):
        ct = t.get("close_time")
        # Always return an ISO string — None and non-datetime values go to the END
        if ct is None:
            return datetime.max.isoformat()
        if isinstance(ct, datetime):
            return ct.isoformat()
        # Fallback: str(ct) for ISO strings already in cache; datetime.max for anything else
        s = str(ct)
        return s if s[:4].isdigit() else datetime.max.isoformat()

    return sorted(merged, key=sort_key)
